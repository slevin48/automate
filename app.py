import streamlit as st
import pandas as pd
import openpyxl as xl
import shutil

st.title("Excel sheet spliter")

st.write("Choose our Template")
with open("pivot.xlsx", "rb") as file:
    st.download_button("Template Excel file", data = file, file_name="pivot.xlsx",mime="application/octet-stream")

st.write("or")

uploaded_file = st.file_uploader("Choose a XLSX file", type="xlsx")

# st.write(uploaded_file)

if uploaded_file is not None:
    df = pd.read_excel(uploaded_file)
    # st.dataframe(df)
    # st.table(df)
    # st.write(uploaded_file)

    workbook = xl.load_workbook(uploaded_file)
    # sheet_1 = workbook['Overview']
    # st.write(workbook.sheetnames)
    sheet = st.selectbox("Select tab", workbook.sheetnames)

    filename = uploaded_file.name.replace('.xlsx','')+'_'+sheet+'.xlsx'
    with open(filename,"wb") as file:
        file.write(uploaded_file.getvalue())

    sn = workbook.sheetnames
    sn.remove(sheet)

    wb = xl.load_workbook(filename)
    for s in sn:
        std = wb[s]
        wb.remove(std)

    wb.save(filename)

    try:
        df2 = pd.read_excel(filename)
        st.write("Preview of "+sheet)
        st.dataframe(df2)
    except:
        st.write("Cannot display table")
    
    with open(filename, "rb") as file:
        st.download_button("Download new Excel file", data = file, file_name= filename ,mime="application/octet-stream")